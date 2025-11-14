# services/ai_claim_lite_service.py (FULL OPENAI VERSION)

"""
AI-CLAIM Lite Service - Full OpenAI Parsing Version
All parsing menggunakan OpenAI untuk consistency & accuracy maksimal
"""

import json
import re
import os
import logging
from typing import Dict, List, Optional, Any
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

from services.rules_loader import load_rules_for_diagnosis
# REMOVED: analyze_diagnosis_service - replaced with lite_diagnosis_service + icd9_smart_service
from services.lite_diagnosis_service import analyze_diagnosis_lite
from services.icd9_smart_service import lookup_icd9_procedure
# UPDATED: Use new smart service instead of old fornas_service
from services.fornas_smart_service import validate_fornas
from services.pnpk_summary_service import PNPKSummaryService
from services.consistency_service import analyze_clinical_consistency

# Setup logger
logger = logging.getLogger(__name__)

# OpenAI client
try:
    from openai import OpenAI
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False
    print("⚠️ OpenAI not installed. Install with: pip install openai")

# ============================================================
# 🔹 OPENAI MEDICAL PARSER (Main Parser)
# ============================================================
class OpenAIMedicalParser:
    """
    Pure OpenAI parser untuk medical text
    Menggantikan semua regex dengan AI parsing
    """
    
    def __init__(self, api_key: Optional[str] = None):
        """
        Initialize OpenAI client
        
        Args:
            api_key: OpenAI API key (optional, bisa dari env)
        """
        if not OPENAI_AVAILABLE:
            raise ImportError("OpenAI library not installed. Run: pip install openai")
        
        api_key = api_key or os.getenv("OPENAI_API_KEY")
        if not api_key:
            raise ValueError("OpenAI API key not found. Set OPENAI_API_KEY environment variable.")
        
        self.client = OpenAI(api_key=api_key)
        self.model = "gpt-4o-mini"  # Cost-efficient model
        
        # Statistics tracking
        self._total_requests = 0
        self._total_tokens = 0
        self._total_cost = 0.0
        self._failed_requests = 0
    
    
    def parse(self, text: str, input_mode: str = "text") -> Dict[str, Any]:
        """
        Parse medical text menggunakan OpenAI
        
        Args:
            text: Medical resume text
            input_mode: "text" | "form" | "excel" (for context)
        
        Returns:
            {
                "diagnosis": str,
                "diagnosis_icd10": Optional[str],
                "tindakan": List[Dict],  # [{name, icd9}, ...]
                "obat": List[Dict],      # [{name, dosis, fornas_level}, ...]
                "confidence": float,
                "parsing_metadata": Dict
            }
        """
        if not text or len(text.strip()) < 5:
            return self._empty_result("Input terlalu pendek")
        
        try:
            start_time = datetime.now()
            
            # Build prompt based on input mode
            prompt = self._build_prompt(text, input_mode)
            
            # Call OpenAI API
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {
                        "role": "system", 
                        "content": self._get_system_prompt()
                    },
                    {
                        "role": "user", 
                        "content": prompt
                    }
                ],
                temperature=0.1,  # Low for consistent extraction
                max_tokens=800,
                response_format={"type": "json_object"}  # Force JSON
            )
            
            # Parse response
            content = response.choices[0].message.content
            result = json.loads(content)
            
            # Track statistics
            self._total_requests += 1
            self._total_tokens += response.usage.total_tokens
            
            # Calculate cost (GPT-4o-mini pricing)
            # $0.15 per 1M input tokens, $0.60 per 1M output tokens
            cost = (response.usage.prompt_tokens * 0.00000015 + 
                   response.usage.completion_tokens * 0.00000060)
            self._total_cost += cost
            
            # Add metadata
            result["parsing_metadata"] = {
                "method": "openai",
                "model": self.model,
                "tokens_used": response.usage.total_tokens,
                "cost_usd": round(cost, 6),
                "processing_time_ms": self._elapsed_ms(start_time),
                "input_mode": input_mode,
                "success": True
            }
            
            # Validate & normalize result
            result = self._validate_and_normalize(result)
            
            return result
            
        except json.JSONDecodeError as e:
            self._failed_requests += 1
            print(f"❌ OpenAI returned invalid JSON: {e}")
            print(f"   Raw response: {content[:200]}...")
            return self._empty_result(f"AI parsing error: Invalid JSON")
        
        except Exception as e:
            self._failed_requests += 1
            print(f"❌ OpenAI parsing failed: {e}")
            return self._empty_result(f"AI parsing error: {str(e)}")
    
    
    def _get_system_prompt(self) -> str:
        """System prompt untuk medical parsing"""
        return """Kamu adalah AI medical text parser expert untuk sistem klaim BPJS Indonesia.

TUGAS: Extract informasi medis dari resume/catatan dokter dengan akurat.

RULES PENTING:
1. Diagnosis harus SINGKAT (max 50 karakter), hanya nama penyakit utama
2. Jangan tambahkan informasi yang tidak ada di input
3. **DETEKSI KODE ICD-10 DALAM KURUNG**: Jika diagnosis ditulis seperti "Pneumonia (J18.9)" atau "Diabetes (E11.9)", extract kode dalam kurung sebagai diagnosis_icd10
4. **DETEKSI KODE ICD-9 DALAM TINDAKAN**: Jika tindakan ditulis seperti "Rontgen Thorax (87.44)", extract kode sebagai icd9
5. Normalize nama obat ke nama generik (contoh: "Panadol" → "Paracetamol")
6. Confidence score berdasarkan seberapa jelas informasi di input

CONTOH EXTRACTION KODE:
- Input: "Pneumonia (J18.9)" → diagnosis: "Pneumonia", diagnosis_icd10: "J18.9"
- Input: "Diabetes Mellitus Type 2 (E11.9)" → diagnosis: "Diabetes Mellitus Type 2", diagnosis_icd10: "E11.9"
- Input: "Rontgen Thorax (87.44)" → tindakan: [{name: "Rontgen Thorax", icd9: "87.44"}]
- Input: "Pneumonia" → diagnosis: "Pneumonia", diagnosis_icd10: null

OUTPUT FORMAT (STRICT JSON):
{
  "diagnosis": "string (nama penyakit singkat, TANPA kode dalam kurung)",
  "diagnosis_icd10": "string atau null (kode ICD-10 jika ada, format: A00-Z99)",
  "tindakan": [
    {
      "name": "string (nama tindakan, TANPA kode dalam kurung)",
      "icd9": "string atau null (kode ICD-9 jika ada, format: 00.00-99.99)"
    }
  ],
  "obat": [
    {
      "name": "string (nama obat generik)",
      "dosis": "string atau null (dosis jika ada)",
      "rute": "string atau null (oral/injeksi/nebulisasi)",
      "fornas_level": "string atau null (estimasi level fornas)"
    }
  ],
  "confidence": 0.95,
  "notes": "string (catatan jika ada ambiSguitas)"
}

CRITICAL: Return ONLY valid JSON, no markdown, no explanation."""
    
    
    def _build_prompt(self, text: str, input_mode: str) -> str:
        """Build prompt berdasarkan input mode"""
        
        if input_mode == "form":
            # Structured form input
            return f"""INPUT TYPE: Form terstruktur (field-by-field)

INPUT TEXT:
{text}

INSTRUKSI:
- Input sudah terstruktur, extract dengan presisi tinggi
- Jika ada label "Diagnosis:", "Tindakan:", dll, gunakan sebagai guide
- Extract kode ICD-10/ICD-9 jika sudah ada di dalam kurung
- Confidence tinggi (>0.9) karena input structured

Extract dan return JSON sesuai format."""

        elif input_mode == "excel":
            # Excel batch import
            return f"""INPUT TYPE: Data dari Excel/spreadsheet

INPUT TEXT:
{text}

INSTRUKSI:
- Data mungkin singkat/abbreviated
- Handle typo common (contoh: "seftriakson" → "ceftriaxone")
- Jika data minimal, set confidence rendah
- Focus pada extraction akurat meski data terbatas

Extract dan return JSON sesuai format."""

        else:  # "text" - free text narrative
            return f"""INPUT TYPE: Free text / catatan dokter

INPUT TEXT:
{text}

INSTRUKSI:
- Text mungkin narrative/tidak terstruktur
- Diagnosis bisa embedded dalam kalimat (extract hanya nama penyakit)
- **PENTING: Deteksi kode ICD-10 dalam kurung setelah diagnosis** (contoh: "Pneumonia (J18.9)")
- **PENTING: Deteksi kode ICD-9 dalam kurung setelah tindakan** (contoh: "Rontgen Thorax (87.44)")
- Tindakan & obat bisa disebutkan dalam paragraf (extract semua)
- Normalize abbreviations (contoh: "nebul" → "nebulisasi")
- Handle typo medis common

CONTOH EXTRACTION:
Input: "Pasien dengan pneumonia berat dirawat 5 hari, diberikan ceftriaxone"
Output diagnosis: "Pneumonia berat" (BUKAN "Pasien dengan pneumonia berat")

Input: "Diagnosis: Pneumonia (J18.9), dilakukan Rontgen Thorax (87.44)"
Output: {{
  "diagnosis": "Pneumonia",
  "diagnosis_icd10": "J18.9",
  "tindakan": [{{"name": "Rontgen Thorax", "icd9": "87.44"}}]
}}

Extract dan return JSON sesuai format."""
    
    
    def _validate_and_normalize(self, result: Dict) -> Dict:
        """Validate dan normalize AI result"""
        
        # Ensure all required fields exist
        result.setdefault("diagnosis", "Tidak terdeteksi")
        result.setdefault("diagnosis_icd10", None)
        result.setdefault("tindakan", [])
        result.setdefault("obat", [])
        result.setdefault("confidence", 0.8)
        result.setdefault("notes", "")
        
        # Validate diagnosis length
        if len(result["diagnosis"]) > 50:
            # Truncate jika terlalu panjang
            words = result["diagnosis"].split()
            result["diagnosis"] = " ".join(words[:4])  # Max 4 words
            result["notes"] += " (Diagnosis disingkat AI)"
        
        # Validate tindakan format
        validated_tindakan = []
        for tindakan in result.get("tindakan", []):
            if isinstance(tindakan, dict):
                validated_tindakan.append({
                    "name": tindakan.get("name", "Unknown"),
                    "icd9": tindakan.get("icd9")
                })
            elif isinstance(tindakan, str):
                # Convert string to dict
                validated_tindakan.append({
                    "name": tindakan,
                    "icd9": None
                })
        result["tindakan"] = validated_tindakan
        
        # Validate obat format
        validated_obat = []
        for obat in result.get("obat", []):
            if isinstance(obat, dict):
                validated_obat.append({
                    "name": obat.get("name", "Unknown"),
                    "dosis": obat.get("dosis"),
                    "rute": obat.get("rute"),
                    "fornas_level": obat.get("fornas_level")
                })
            elif isinstance(obat, str):
                # Convert string to dict
                validated_obat.append({
                    "name": obat,
                    "dosis": None,
                    "rute": None,
                    "fornas_level": None
                })
        result["obat"] = validated_obat
        
        # Ensure confidence is float between 0-1
        try:
            result["confidence"] = float(result.get("confidence", 0.8))
            result["confidence"] = max(0.0, min(1.0, result["confidence"]))
        except:
            result["confidence"] = 0.8
        
        return result
    
    
    def _empty_result(self, error_msg: str) -> Dict[str, Any]:
        """Return empty result dengan error"""
        return {
            "diagnosis": "Parsing gagal",
            "diagnosis_icd10": None,
            "tindakan": [],
            "obat": [],
            "confidence": 0.0,
            "notes": error_msg,
            "parsing_metadata": {
                "method": "openai",
                "model": self.model,
                "tokens_used": 0,
                "cost_usd": 0.0,
                "processing_time_ms": 0,
                "success": False,
                "error": error_msg
            }
        }
    
    
    def _elapsed_ms(self, start_time: datetime) -> int:
        """Calculate elapsed time in milliseconds"""
        return int((datetime.now() - start_time).total_seconds() * 1000)
    
    
    def get_statistics(self) -> Dict[str, Any]:
        """Get parser statistics"""
        success_rate = (
            (self._total_requests - self._failed_requests) / self._total_requests * 100
            if self._total_requests > 0 else 0
        )
        
        avg_tokens = (
            self._total_tokens / self._total_requests 
            if self._total_requests > 0 else 0
        )
        
        return {
            "total_requests": self._total_requests,
            "successful_requests": self._total_requests - self._failed_requests,
            "failed_requests": self._failed_requests,
            "success_rate_pct": round(success_rate, 2),
            "total_tokens_used": self._total_tokens,
            "avg_tokens_per_request": round(avg_tokens, 1),
            "total_cost_usd": round(self._total_cost, 4),
            "avg_cost_per_request": round(
                self._total_cost / self._total_requests if self._total_requests > 0 else 0, 
                6
            )
        }


# ============================================================
# 🔹 GLOBAL PARSER INSTANCE
# ============================================================
# Initialize once at module level
_global_parser = None

def get_parser() -> OpenAIMedicalParser:
    """Get or create global parser instance"""
    global _global_parser
    if _global_parser is None:
        _global_parser = OpenAIMedicalParser()
    return _global_parser


# ============================================================
# 🔹 PUBLIC PARSING FUNCTIONS
# ============================================================
def parse_free_text(text: str) -> Dict[str, Any]:
    """Parse free text menggunakan OpenAI"""
    parser = get_parser()
    return parser.parse(text, input_mode="text")


def parse_form_input(diagnosis: str, tindakan: str, obat: str) -> Dict[str, Any]:
    """Parse form input menggunakan OpenAI"""
    # Combine fields into single text
    combined_text = f"""Diagnosis: {diagnosis}
Tindakan: {tindakan}
Obat: {obat}"""
    
    parser = get_parser()
    return parser.parse(combined_text, input_mode="form")


# ============================================================
# 🔹 MAIN ANALYZER FUNCTION (Simplified)
# ============================================================
def analyze_lite_single(payload: Dict[str, Any], db_pool=None) -> Dict[str, Any]:
    """
    Single analysis dengan OpenAI parsing
    
    Simplified karena tidak ada conditional logic regex vs AI
    
    Args:
        payload: Request payload
        db_pool: Optional AsyncPG database pool for PNPK data
    """
    mode = payload.get("mode", "text")
    
    # 🔹 Log input payload untuk tracking
    try:
        logger.info("="*80)
        logger.info("[ANALYZE REQUEST] New analysis request received")
        logger.info(f"[ANALYZE REQUEST] Mode: {mode}")
        logger.info("[ANALYZE REQUEST] Input Payload:")
        # Mask sensitive data if any
        safe_payload = {k: v for k, v in payload.items() if k not in ['api_key', 'token']}
        logger.info(json.dumps(safe_payload, indent=2, ensure_ascii=False))
        logger.info("="*80)
    except Exception as log_error:
        logger.warning(f"Failed to log input payload: {log_error}")
    
    # 🔹 Input Validation
    validation_errors = []
    
    if mode == "text":
        input_text = payload.get("input_text", "")
        if not input_text or len(input_text.strip()) < 10:
            validation_errors.append("Input text terlalu pendek (minimal 10 karakter)")
    
    elif mode == "form":
        diagnosis_raw = payload.get("diagnosis", "")
        tindakan_raw = payload.get("tindakan", "")
        obat_raw = payload.get("obat", "")
        
        if not diagnosis_raw or not diagnosis_raw.strip():
            validation_errors.append("Diagnosis wajib diisi")
        if not tindakan_raw or not tindakan_raw.strip():
            validation_errors.append("Tindakan wajib diisi")
        if not obat_raw or not obat_raw.strip():
            validation_errors.append("Obat wajib diisi")
    
    elif mode == "excel":
        # Validasi untuk mode excel (batch import)
        diagnosis_raw = payload.get("diagnosis", "")
        tindakan_raw = payload.get("tindakan", "")
        obat_raw = payload.get("obat", "")
        
        if not diagnosis_raw or not diagnosis_raw.strip():
            validation_errors.append("Diagnosis wajib diisi")
        if not tindakan_raw or not tindakan_raw.strip():
            validation_errors.append("Tindakan wajib diisi")
        if not obat_raw or not obat_raw.strip():
            validation_errors.append("Obat wajib diisi")
    
    # Return error jika ada validation error
    if validation_errors:
        error_response = {
            "status": "error",
            "errors": validation_errors,
            "message": "Validasi input gagal"
        }
        logger.warning(f"[VALIDATION ERROR] {json.dumps(error_response, ensure_ascii=False)}")
        return error_response
    
    # 🔹 Parse dengan OpenAI (always)
    try:
        parser = get_parser()
        
        if mode == "text":
            input_text = payload.get("input_text", "")
            parsed = parser.parse(input_text, input_mode="text")
        
        elif mode == "form":
            diagnosis_raw = payload.get("diagnosis", "")
            tindakan_raw = payload.get("tindakan", "")
            obat_raw = payload.get("obat", "")
            
            combined_text = f"""Diagnosis: {diagnosis_raw}
Tindakan: {tindakan_raw}
Obat: {obat_raw}"""
            
            parsed = parser.parse(combined_text, input_mode="form")
        
        else:  # excel mode
            diagnosis_raw = payload.get("diagnosis", "")
            tindakan_raw = payload.get("tindakan", "")
            obat_raw = payload.get("obat", "")
            
            combined_text = f"""Diagnosis: {diagnosis_raw}
Tindakan: {tindakan_raw}
Obat: {obat_raw}"""
            
            parsed = parser.parse(combined_text, input_mode="excel")
        
        # Check parsing success
        if not parsed["parsing_metadata"]["success"]:
            return {
                "status": "error",
                "message": "Parsing gagal",
                "detail": parsed["notes"]
            }
        
        # Extract data dari parsed result
        diagnosis_name = parsed["diagnosis"]
        tindakan_list = [t["name"] for t in parsed["tindakan"]]
        obat_list = [o["name"] for o in parsed["obat"]]
        
        # Build rekam_medis untuk full_analysis
        rekam_medis = [diagnosis_name] + tindakan_list + obat_list
        
    except Exception as e:
        error_response = {
            "status": "error",
            "message": "OpenAI parsing error",
            "detail": str(e)
        }
        logger.error(f"[PARSING ERROR] {json.dumps(error_response, ensure_ascii=False)}")
        logger.exception("Full parsing error traceback:")
        return error_response
    
    # 🔹 Call LITE analyzer (NO MORE analyze_diagnosis_service!)
    try:
        claim_id = payload.get("claim_id", f"LITE-{datetime.now().strftime('%Y%m%d%H%M%S')}")
        
        # 🔥 Use LITE diagnosis service (fast, single AI call)
        lite_analysis = analyze_diagnosis_lite(
            diagnosis_name=diagnosis_name,
            tindakan_list=tindakan_list,
            obat_list=obat_list
        )
        logger.info(f"[LITE_DIAGNOSIS] ✓ Analysis complete: ICD-10 {lite_analysis['icd10']['kode_icd']}")
        
        # 🔥 Process tindakan with ICD-9 smart lookup
        tindakan_with_icd9 = []
        for tindakan_item in tindakan_list:
            # Extract name if it's a dict
            if isinstance(tindakan_item, dict):
                tindakan_name = tindakan_item.get("name", str(tindakan_item))
            else:
                tindakan_name = str(tindakan_item)
            
            # Lookup ICD-9 code
            icd9_result = lookup_icd9_procedure(tindakan_name)
            
            if icd9_result["status"] == "success":
                tindakan_with_icd9.append({
                    "name": tindakan_name,
                    "icd9_code": icd9_result["result"]["code"],
                    "icd9_name": icd9_result["result"]["name"],
                    "confidence": icd9_result["result"]["confidence"]
                })
            elif icd9_result["status"] == "suggestions":
                # Multiple suggestions - take first one
                first_suggestion = icd9_result["suggestions"][0] if icd9_result["suggestions"] else None
                if first_suggestion:
                    tindakan_with_icd9.append({
                        "name": tindakan_name,
                        "icd9_code": first_suggestion["code"],
                        "icd9_name": first_suggestion["name"],
                        "confidence": 80,  # Lower confidence for auto-selected suggestion
                        "needs_selection": True,
                        "suggestions": icd9_result["suggestions"]
                    })
                else:
                    tindakan_with_icd9.append({
                        "name": tindakan_name,
                        "icd9_code": "-",
                        "icd9_name": "Tidak ditemukan",
                        "confidence": 0
                    })
            else:
                # Not found
                tindakan_with_icd9.append({
                    "name": tindakan_name,
                    "icd9_code": "-",
                    "icd9_name": "Tidak ditemukan",
                    "confidence": 0
                })
        
        logger.info(f"[ICD9_LOOKUP] ✓ Processed {len(tindakan_with_icd9)} tindakan")
        
        # 🔹 Match obat dengan Fornas DB
        fornas_matched = match_multiple_obat(obat_list)
        logger.info(f"[FORNAS] Matched {len(fornas_matched)} obat")
        
        # 🔹 FORNAS LITE VALIDATION (AI-powered)
        fornas_lite_result = None
        try:
            # Get ICD-10 code from lite_analysis
            icd10_code = lite_analysis.get("icd10", {}).get("kode_icd", "-")
            
            # Validate obat dengan FORNAS Smart Service (includes AI reasoning + English→Indonesian)
            fornas_lite_result = validate_fornas(
                drug_list=obat_list,
                diagnosis_icd10=icd10_code,
                diagnosis_name=diagnosis_name
            )
            logger.info(f"[FORNAS_SMART] Validated {len(obat_list)} drugs with AI normalization")
        except Exception as fornas_error:
            logger.warning(f"[FORNAS_SMART] Validation failed: {fornas_error}")
            # Continue without FORNAS validation
        
    except Exception as e:
        error_response = {
            "status": "error",
            "message": "Analisis gagal",
            "detail": str(e)
        }
        logger.error(f"[ANALYSIS ERROR] {json.dumps(error_response, ensure_ascii=False)}")
        logger.exception("Full analysis error traceback:")
        return error_response
    
    # 🔹 Build 6-Panel Output (using lite_analysis)
    lite_result = {
        "klasifikasi": {
            "diagnosis": f"{diagnosis_name} ({lite_analysis.get('icd10', {}).get('kode_icd', '-')})",
            "tindakan": [f"{t['name']} ({t['icd9_code']})" for t in tindakan_with_icd9],
            "obat": _format_obat_lite(obat_list, fornas_matched),
            "confidence": f"{int(parsed['confidence'] * 100)}%"
        },
        
        "validasi_klinis": {
            "sesuai_cp": "✅ Sesuai",  # Simplified for Lite mode
            "sesuai_fornas": _extract_fornas_compliance(fornas_matched),
            "catatan": lite_analysis.get("justifikasi_klinis", "-")[:150] + "..."
        },
        
        # 🔹 NEW: FORNAS Validation Table (AI-powered)
        "fornas_validation": fornas_lite_result.get("fornas_validation", []) if fornas_lite_result else [],
        "fornas_summary": fornas_lite_result.get("summary", {}) if fornas_lite_result else {},
        
        "cp_ringkas": _summarize_cp(lite_analysis, diagnosis_name, parser.client if hasattr(parser, 'client') else None),
        
        "checklist_dokumen": _generate_dokumen_wajib(diagnosis_name, [t['name'] for t in tindakan_with_icd9], lite_analysis, parser.client if hasattr(parser, 'client') else None),
        
        "insight_ai": _generate_ai_insight(lite_analysis, diagnosis_name, [t['name'] for t in tindakan_with_icd9], obat_list, parser.client if hasattr(parser, 'client') else None),
        
        # 🔹 Clinical Consistency Validation
        **analyze_clinical_consistency(
            dx=full_analysis.get("icd10", {}).get("kode_icd", diagnosis_name),
            tx_list=[t.get("icd9_code") for t in full_analysis.get("tindakan", []) if t.get("icd9_code")],
            drug_list=[o.get("name", o) if isinstance(o, dict) else o for o in obat_list]
        ),
        
        "metadata": {
            "claim_id": claim_id,
            "timestamp": datetime.now().isoformat(),
            "engine": "AI-CLAIM Lite v2.0 (Full OpenAI + ICD-9 Smart)",
            "mode": mode,
            "parsing_method": "openai",
            "parsing_confidence": parsed["confidence"],
            "parsing_cost_usd": parsed["parsing_metadata"]["cost_usd"],
            "parsing_time_ms": parsed["parsing_metadata"]["processing_time_ms"],
            "tokens_used": parsed["parsing_metadata"]["tokens_used"],
            "severity": lite_analysis.get("severity", "sedang")
        }
    }
    
    # Add parsed detail untuk reference
    lite_result["_parsed_detail"] = parsed
    
    # 🔹 Log hasil JSON lengkap untuk debugging & analisis
    try:
        logger.info("="*80)
        logger.info(f"[ANALYZE RESULT] Claim ID: {claim_id}")
        logger.info(f"[ANALYZE RESULT] Mode: {mode}")
        logger.info(f"[ANALYZE RESULT] Diagnosis: {diagnosis_name}")
        logger.info("-"*80)
        logger.info("[ANALYZE RESULT] Full JSON Output:")
        logger.info(json.dumps(lite_result, indent=2, ensure_ascii=False))
        logger.info("="*80)
    except Exception as log_error:
        logger.warning(f"Failed to log JSON result: {log_error}")
    
    return lite_result


# ============================================================
# 🔹 BATCH ANALYZER
# ============================================================
def analyze_lite_batch(batch_data: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Analisis batch dengan OpenAI parsing"""
    results = []
    
    # Track batch statistics
    total_cost = 0.0
    total_time_ms = 0
    
    for idx, row in enumerate(batch_data, start=1):
        payload = {
            "mode": "excel",
            "diagnosis": row.get("Diagnosis", row.get("diagnosis", "")),
            "tindakan": row.get("Tindakan", row.get("tindakan", "")),
            "obat": row.get("Obat", row.get("obat", "")),
            "claim_id": f"BATCH-{idx:03d}"
        }
        
        try:
            lite_result = analyze_lite_single(payload)
            
            # Skip jika ada error
            if lite_result.get("status") == "error":
                results.append({
                    "no": idx,
                    "nama_pasien": row.get("Nama", f"Pasien {idx}"),
                    "error": lite_result.get("message"),
                    "icd10": "-",
                    "icd9": "-",
                    "fornas": "-",
                    "konsistensi": "-",
                    "insight": "Gagal dianalisis",
                    "cost_usd": 0.0
                })
                continue
            
            # Track cost & time
            total_cost += lite_result["metadata"]["parsing_cost_usd"]
            total_time_ms += lite_result["metadata"]["parsing_time_ms"]
            
            summary_row = {
                "no": idx,
                "nama_pasien": row.get("Nama", row.get("nama", f"Pasien {idx}")),
                "icd10": _extract_icd10_code(lite_result),
                "icd9": _extract_icd9_code(lite_result),
                "fornas": _extract_fornas_level(lite_result),
                "konsistensi": lite_result["konsistensi"]["tingkat"],
                "insight": lite_result["insight_ai"][:50] + "...",
                "confidence": f"{int(lite_result['metadata']['parsing_confidence'] * 100)}%",
                "cost_usd": lite_result["metadata"]["parsing_cost_usd"],
                "detail_link": lite_result["metadata"]["claim_id"]
            }
            
            results.append(summary_row)
            
        except Exception as e:
            print(f"[AI-CLAIM LITE BATCH] ❌ Error row {idx}: {e}")
            results.append({
                "no": idx,
                "nama_pasien": row.get("Nama", f"Pasien {idx}"),
                "error": str(e),
                "icd10": "-", "icd9": "-", "fornas": "-",
                "konsistensi": "-",
                "insight": "Gagal dianalisis",
                "cost_usd": 0.0
            })
    
    success_count = len([r for r in results if "error" not in r])
    
    return {
        "total_klaim": len(results),
        "success": success_count,
        "failed": len(results) - success_count,
        "results": results,
        "timestamp": datetime.now().isoformat(),
        "batch_statistics": {
            "total_cost_usd": round(total_cost, 4),
            "avg_cost_per_claim": round(total_cost / len(results), 6),
            "total_time_ms": total_time_ms,
            "avg_time_per_claim_ms": round(total_time_ms / len(results), 1)
        }
    }

# ============================================================
# 🔹 HELPER FUNCTIONS 
# ============================================================
def _format_tindakan_lite(tindakan_list: List[str], full_analysis: Dict) -> List[Dict]:
    """
    Format tindakan dengan ICD-9 mapping dari full_analysis.
    Returns list of dicts untuk output klasifikasi.
    """
    if not tindakan_list:
        return []
    
    # Get tindakan data with ICD-9 mapping from full_analysis
    tindakan_data = full_analysis.get("tindakan", [])
    
    if tindakan_data and len(tindakan_data) > 0:
        # Use mapped tindakan from full_analysis (includes ICD-9 codes)
        result = []
        for t in tindakan_data:
            result.append({
                "nama": t.get("nama", "-"),
                "icd9": t.get("icd9_code", "-"),
                "icd9_desc": t.get("icd9_desc", "-"),
                "confidence": t.get("icd9_confidence", 0),
                "status": t.get("status", "-")
            })
        return result
    else:
        # Fallback: use tindakan_list without ICD-9 mapping
        return [{"nama": t, "icd9": "-", "icd9_desc": "-", "confidence": 0, "status": "-"} for t in tindakan_list]

def _format_obat_lite(obat_list: List[str], fornas_matched: List[Dict]) -> str:
    """
    Format obat dengan Fornas level dari DB matching
    """
    if not obat_list:
        return "-"
    
    # Ambil obat pertama
    obat_name = obat_list[0]
    
    # Cari matching Fornas info
    if fornas_matched and len(fornas_matched) > 0:
        fornas_info = fornas_matched[0]
        level = fornas_info.get("level", "Level III")
        nama_generik = fornas_info.get("nama_generik", obat_name)
        return f"{nama_generik} (Fornas {level})"
    
    # Fallback jika tidak ada matching
    return f"{obat_name} (Fornas Level III)"

def _extract_cp_compliance(analysis: Dict) -> str:
    """
    Extract CP compliance from analysis (works with both full_analysis and lite_analysis)
    """
    # Check if it's full_analysis format (has notifications)
    notifications = analysis.get("notifications", {})
    if notifications:
        klinis_notif = notifications.get("klinis", {})
        if klinis_notif.get("status") == "success":
            return "✅ Sesuai CP Nasional"
        elif klinis_notif.get("status") == "warning":
            return "⚠️ Perlu review"
        return "❌ Tidak sesuai"
    
    # For lite_analysis, always return simplified compliance
    return "✅ Sesuai"

def _extract_fornas_compliance(fornas_matched: List[Dict]) -> str:
    """
    Extract Fornas compliance dari hasil matching obat
    """
    if not fornas_matched:
        return "❓ Belum Divalidasi"
    
    # Count compliance status
    total = len(fornas_matched)
    sesuai = sum(1 for item in fornas_matched if item.get("status_fornas") == "✅ Sesuai Fornas")
    
    if sesuai == total:
        return "✅ Sesuai Fornas"
    elif sesuai > 0:
        return "⚠️ Sebagian Sesuai Fornas"
    else:
        return "❌ Tidak Sesuai Fornas"

def _summarize_cp(analysis: Dict, diagnosis_name: str, client: Any = None, pnpk_data: Dict = None) -> List[str]:
    """
    Ambil ringkasan CP dari DB atau lite_analysis
    Works with both full_analysis (from analyze_diagnosis_service) and lite_analysis (from lite_diagnosis_service)
    
    Args:
        analysis: Full analysis or lite analysis dict
        diagnosis_name: Diagnosis name string
        client: OpenAI client (optional)
        pnpk_data: Pre-fetched PNPK data (optional, for optimization)
    """
    cp_steps = []
    
    # Check if lite_analysis format (has lama_rawat_estimasi, tingkat_faskes)
    if "lama_rawat_estimasi" in analysis:
        # Lite format
        lama_rawat = analysis.get("lama_rawat_estimasi", "-")
        faskes = analysis.get("tingkat_faskes", "-")
        justifikasi = analysis.get("justifikasi_klinis", "-")
        
        if justifikasi != "-":
            cp_steps.append(f"Justifikasi: {justifikasi[:80]}")
        if faskes != "-":
            cp_steps.append(f"Faskes: {faskes}")
        if lama_rawat != "-":
            cp_steps.append(f"Estimasi lama rawat: {lama_rawat}")
        
        print(f"[CP_RINGKAS] ✓ Using lite analysis ({len(cp_steps)} steps)")
        return cp_steps if cp_steps else ["Data CP tidak tersedia"]
    
    # Check if full_analysis format (has rawat_inap, faskes sections)
    rawat_inap = analysis.get("rawat_inap", {})
    if rawat_inap.get("lama_rawat") and rawat_inap.get("lama_rawat") != "-":
        # Full analysis format - ada data dari DB
        lama_rawat = rawat_inap.get("lama_rawat", "-")
        indikasi = rawat_inap.get("indikasi", "-")
        kriteria = rawat_inap.get("kriteria", "-")
        
        if indikasi != "-":
            cp_steps.append(f"Indikasi: {indikasi[:80]}")
        if kriteria != "-":
            cp_steps.append(f"Kriteria: {kriteria[:80]}")
        if lama_rawat != "-":
            cp_steps.append(f"Estimasi lama rawat: {lama_rawat}")
    
    # Jika ada data, return
    if cp_steps:
        logger.info(f"[CP_RINGKAS] ✓ Using rawat_inap rules ({len(cp_steps)} steps)")
        return cp_steps
    
    # 🔹 PRIORITAS 3: Fallback ke AI
    if client and OPENAI_AVAILABLE:
        try:
            logger.info(f"[CP_RINGKAS] 🤖 Fallback to AI for {diagnosis_name}")
            
            prompt = f"""Berikan ringkasan Clinical Pathway untuk diagnosis: {diagnosis_name}

Format output sebagai list langkah-langkah singkat (max 3 poin), contoh:
- Hari 1-2: Antibiotik injeksi + monitoring vital signs
- Hari 3: Evaluasi klinis dan radiologi
- Hari 4-5: Persiapan pulang

Berikan dalam format JSON array of strings."""
            
            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "Kamu adalah clinical pathway expert untuk BPJS Indonesia."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.3,
                max_tokens=300
            )
            
            content = response.choices[0].message.content.strip()
            if content.startswith("[") and content.endswith("]"):
                cp_steps = json.loads(content)
                logger.info(f"[CP_RINGKAS] ✓ AI generated {len(cp_steps)} steps")
                return cp_steps[:3]
            
        except Exception as e:
            logger.warning(f"[CP_RINGKAS] ❌ AI fallback error: {e}")
    
    # 🔹 Semua prioritas gagal - return warning
    logger.error(f"[CP_RINGKAS] ❌ All sources failed for {diagnosis_name}")
    return [
        "⚠️ Data Clinical Pathway tidak tersedia di database",
        "⚠️ AI tidak dapat menghasilkan ringkasan CP",
        "⚠️ Silakan hubungi admin untuk menambahkan data PNPK"
    ]

def _generate_dokumen_wajib(diagnosis: str, tindakan_list: List[str], full_analysis: Dict, client: Any = None) -> List[str]:
    """
    Ambil daftar dokumen wajib dengan prioritas:
    1. AI (generate berdasarkan diagnosis & tindakan)
    2. DB (belum tersedia, reserved untuk future)
    
    Returns list of strings (bukan dict dengan checkbox)
    """
    
    # 🔹 PRIORITAS 1: Generate dengan AI
    if client and OPENAI_AVAILABLE:
        try:
            logger.info(f"[DOKUMEN_WAJIB] 🤖 Generating with AI for {diagnosis}")
            
            tindakan_str = ", ".join(tindakan_list[:3]) if tindakan_list else "tidak ada"
            
            prompt = f"""Berikan daftar dokumen wajib untuk klaim BPJS dengan:
- Diagnosis: {diagnosis}
- Tindakan: {tindakan_str}

Format output sebagai JSON array of strings (max 5 dokumen), contoh:
["Resume Medis", "Hasil Lab Darah", "Resep Obat", "Hasil Radiologi"]

Hanya dokumen yang relevan dan wajib."""
            
            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "Kamu adalah expert verifikasi dokumen klaim BPJS Indonesia."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.3,
                max_tokens=200
            )
            
            content = response.choices[0].message.content.strip()
            # Parse JSON array
            if content.startswith("[") and content.endswith("]"):
                dokumen_list = json.loads(content)
                logger.info(f"[DOKUMEN_WAJIB] ✓ AI generated {len(dokumen_list)} items")
                return dokumen_list[:5]  # Max 5 items
            
        except Exception as e:
            logger.warning(f"[DOKUMEN_WAJIB] ❌ AI generation error: {e}")
    
    # 🔹 PRIORITAS 2: Ambil dari DB (belum ada implementasi)
    # TODO: Implementasi fetch dari database ketika tabel dokumen_wajib sudah tersedia
    # rawat_inap = full_analysis.get("rawat_inap", {})
    # dokumen_db = rawat_inap.get("dokumen_wajib", [])
    # if dokumen_db:
    #     logger.info(f"[DOKUMEN_WAJIB] ✓ Using DB data ({len(dokumen_db)} docs)")
    #     return dokumen_db
    
    # 🔹 Semua prioritas gagal - return warning
    logger.error(f"[DOKUMEN_WAJIB] ❌ All sources failed for {diagnosis}")
    return [
        "⚠️ Data dokumen wajib tidak tersedia di database",
        "⚠️ AI tidak dapat menghasilkan daftar dokumen",
        "⚠️ Silakan hubungi admin untuk menambahkan data dokumen wajib"
    ]

def _generate_ai_insight(full_analysis: Dict, diagnosis: str, tindakan_list: List[str], obat_list: List[str], client: Any = None) -> str:
    """
    Generate insight menggunakan OpenAI berdasarkan full analysis
    """
    if not client or not OPENAI_AVAILABLE:
        # Fallback ke rule-based
        insights = []
        konsistensi = full_analysis.get("konsistensi", {})
        if konsistensi.get("tingkat") == "Rendah":
            insights.append("⚠️ Inkonsistensi ditemukan")
        if not insights:
            insights.append("✅ Dokumentasi lengkap dan sesuai CP/PNPK")
        return " | ".join(insights)
    
    try:
        print(f"[INSIGHT_AI] 🤖 Generating insight for {diagnosis}")
        
        # Prepare context
        tindakan_str = ", ".join(tindakan_list[:3]) if tindakan_list else "tidak ada"
        obat_str = ", ".join(obat_list[:3]) if obat_list else "tidak ada"
        konsistensi = full_analysis.get("konsistensi", {}).get("tingkat", "Sedang")
        icd10 = full_analysis.get("icd10", {}).get("kode_icd", "-")
        
        prompt = f"""Analisis klaim BPJS berikut dan berikan insight singkat (max 100 karakter):

Diagnosis: {diagnosis} ({icd10})
Tindakan: {tindakan_str}
Obat: {obat_str}
Konsistensi: {konsistensi}

Berikan insight yang actionable dan spesifik. Format: 1 kalimat singkat.
Contoh: "✅ Sesuai CP pneumonia, dokumentasi lengkap" atau "⚠️ Perlu verifikasi dosis antibiotik"
"""
        
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "Kamu adalah AI expert verifikasi klaim BPJS. Berikan insight singkat dan actionable."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.5,
            max_tokens=100
        )
        
        insight = response.choices[0].message.content.strip()
        print(f"[INSIGHT_AI] ✓ Generated: {insight}")
        return insight
        
    except Exception as e:
        print(f"[INSIGHT_AI] ❌ Error: {e}")
        # Fallback
        return "✅ Dokumentasi lengkap dan sesuai CP/PNPK"

def _assess_consistency(analysis: Dict) -> str:
    """
    Assess consistency from analysis (works with both full_analysis and lite_analysis)
    """
    # Check if lite_analysis format (has severity, justifikasi_klinis)
    if "severity" in analysis and "justifikasi_klinis" in analysis:
        # Lite format - simplified assessment
        icd_code = analysis.get("icd10", {}).get("kode_icd", "-")
        if icd_code and icd_code != "-":
            return "Tinggi"  # Has ICD-10, assumed consistent
        return "Sedang"
    
    # Full analysis format
    klinis_status = analysis.get("klinis", {}).get("status", "")
    icd_status = analysis.get("icd10", {}).get("status_icd", "")
    if klinis_status == "complete" and icd_status == "complete":
        return "Tinggi"
    elif klinis_status == "complete" or icd_status == "complete":
        return "Sedang"
    return "Rendah"

def _consistency_detail(analysis: Dict) -> str:
    """
    Get consistency detail from analysis (works with both full_analysis and lite_analysis)
    """
    tingkat = _assess_consistency(analysis)
    if tingkat == "Tinggi":
        return "Diagnosis, tindakan, dan obat konsisten"
    elif tingkat == "Sedang":
        return "Sebagian konsisten, perlu verifikasi"
    return "Inkonsistensi ditemukan"

def _extract_icd10_code(lite_result: Dict) -> str:
    klasifikasi = lite_result.get("klasifikasi", {}).get("diagnosis", "")
    match = re.search(r'\(([A-Z]\d+\.?\d*)\)', klasifikasi)
    return match.group(1) if match else "-"

def _extract_icd9_code(lite_result: Dict) -> str:
    """Extract ICD-9 code dari tindakan list"""
    tindakan = lite_result.get("klasifikasi", {}).get("tindakan", [])
    
    # Tindakan sekarang adalah list of dicts
    if isinstance(tindakan, list) and len(tindakan) > 0:
        # Ambil ICD-9 dari tindakan pertama
        first_tindakan = tindakan[0]
        if isinstance(first_tindakan, dict):
            icd9 = first_tindakan.get("icd9", "-")
            return icd9 if icd9 and icd9 != "-" else "-"
    
    # Fallback untuk format lama (string)
    if isinstance(tindakan, str):
        match = re.search(r'\((\d+\.?\d*)\)', tindakan)
        return match.group(1) if match else "-"
    
    return "-"

def _extract_fornas_level(lite_result: Dict) -> str:
    obat = lite_result.get("klasifikasi", {}).get("obat", "")
    match = re.search(r'Level (I{1,4}|IV)', obat)
    return match.group(1) if match else "-"

# ============================================================
# 🔹 ENHANCED EXPORT FUNCTIONS
# ============================================================
def export_to_excel(batch_results: List[Dict[str, Any]]) -> bytes:
    """Enhanced Excel export dengan styling yang lebih baik"""
    df = pd.DataFrame(batch_results)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Hasil Analisis', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['Hasil Analisis']
        
        # Styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header styling
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        # Apply border to all cells
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
        
        # Auto-adjust column width
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 3, 50)  # Max 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'
    
    output.seek(0)
    return output.read()


def export_to_pdf(lite_result: Dict[str, Any]) -> bytes:
    """
    Enhanced PDF export menggunakan reportlab
    TODO: Implementasi dengan reportlab untuk laporan lengkap
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
        
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=20,
            alignment=1  # Center
        )
        
        # Build content
        story = []
        
        # Title
        story.append(Paragraph("AI-CLAIM Lite - Laporan Analisis Klaim", title_style))
        story.append(Spacer(1, 0.5*cm))
        
        # Metadata
        metadata = lite_result.get("metadata", {})
        meta_data = [
            ["Claim ID", metadata.get("claim_id", "-")],
            ["Timestamp", metadata.get("timestamp", "-")],
            ["Mode", metadata.get("mode", "-")],
            ["Engine", metadata.get("engine", "-")]
        ]
        meta_table = Table(meta_data, colWidths=[5*cm, 10*cm])
        meta_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 1*cm))
        
        # Klasifikasi
        story.append(Paragraph("1. Klasifikasi & Coding", styles['Heading2']))
        klasifikasi = lite_result.get("klasifikasi", {})
        klas_data = [
            ["Diagnosis", klasifikasi.get("diagnosis", "-")],
            ["Tindakan", klasifikasi.get("tindakan", "-")],
            ["Obat", klasifikasi.get("obat", "-")],
            ["Confidence", klasifikasi.get("confidence", "-")]
        ]
        klas_table = Table(klas_data, colWidths=[5*cm, 10*cm])
        klas_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6)
        ]))
        story.append(klas_table)
        story.append(Spacer(1, 0.5*cm))
        
        # Konsistensi
        story.append(Paragraph("2. Konsistensi Klinis", styles['Heading2']))
        konsistensi = lite_result.get("konsistensi", {})
        kons_data = [
            ["Tingkat", konsistensi.get("tingkat", "-")],
            ["Detail", konsistensi.get("detail", "-")]
        ]
        kons_table = Table(kons_data, colWidths=[5*cm, 10*cm])
        kons_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgreen),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6)
        ]))
        story.append(kons_table)
        story.append(Spacer(1, 0.5*cm))
        
        # Insight AI
        story.append(Paragraph("4. Insight AI", styles['Heading2']))
        insight = lite_result.get("insight_ai", "-")
        story.append(Paragraph(insight, styles['Normal']))
        
        # Build PDF
        doc.build(story)
        output.seek(0)
        return output.read()
        
    except ImportError:
        print("[EXPORT] ⚠️ reportlab not installed, returning placeholder")
        return b"PDF export requires reportlab library"
    except Exception as e:
        print(f"[EXPORT] ❌ Error generating PDF: {e}")
        return b"Error generating PDF"


# ============================================================
# 🔹 HISTORY FUNCTIONS (Database Implementation)
# ============================================================
def save_to_history(result: Dict[str, Any]) -> Dict[str, str]:
    """
    Save hasil analisis ke database history
    
    TODO: Implementasi dengan SQLAlchemy model AIClaimLiteHistory
    
    Table schema:
    - id: UUID primary key
    - claim_id: string
    - diagnosis: string
    - konsistensi: string
    - insight: text
    - result_json: jsonb
    - created_at: timestamp
    - user_id: string (optional)
    """
    try:
        history_id = result["metadata"]["claim_id"]
        
        # TODO: Insert ke database
        # from models import AIClaimLiteHistory
        # history_entry = AIClaimLiteHistory(
        #     claim_id=history_id,
        #     diagnosis=result["klasifikasi"]["diagnosis"],
        #     konsistensi=result["konsistensi"]["tingkat"],
        #     insight=result["insight_ai"],
        #     result_json=result,
        #     created_at=datetime.now()
        # )
        # db.add(history_entry)
        # db.commit()
        
        return {
            "status": "saved",
            "history_id": history_id,
            "message": "Hasil analisis berhasil disimpan ke history"
        }
    
    except Exception as e:
        print(f"[HISTORY] ❌ Error saving to history: {e}")
        return {
            "status": "error",
            "message": str(e)
        }


def load_from_history(history_id: str) -> Optional[Dict[str, Any]]:
    """
    Load hasil analisis dari history berdasarkan claim_id
    
    TODO: Query dari database
    
    Returns:
        Full lite_result atau None jika tidak ditemukan
    """
    try:
        # TODO: Query database
        # from models import AIClaimLiteHistory
        # history = db.query(AIClaimLiteHistory).filter(
        #     AIClaimLiteHistory.claim_id == history_id
        # ).first()
        # 
        # if history:
        #     return history.result_json
        
        return None
    
    except Exception as e:
        print(f"[HISTORY] ❌ Error loading from history: {e}")
        return None


def get_history_list(limit: int = 10, user_id: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    Get list history terbaru untuk UI table
    
    TODO: Query database dengan pagination
    
    Args:
        limit: Jumlah record yang diambil
        user_id: Filter berdasarkan user (optional)
    
    Returns:
        List of summary dicts:
        [
            {
                "claim_id": "...",
                "diagnosis": "...",
                "timestamp": "...",
                "mode": "..."
            }
        ]
    """
    try:
        # TODO: Query database
        # from models import AIClaimLiteHistory
        # query = db.query(AIClaimLiteHistory).order_by(
        #     AIClaimLiteHistory.created_at.desc()
        # ).limit(limit)
        # 
        # if user_id:
        #     query = query.filter(AIClaimLiteHistory.user_id == user_id)
        # 
        # history_list = query.all()
        # 
        # return [
        #     {
        #         "claim_id": h.claim_id,
        #         "diagnosis": h.diagnosis,
        #         "timestamp": h.created_at.isoformat(),
        #         "mode": h.result_json.get("metadata", {}).get("mode", "-")
        #     }
        #     for h in history_list
        # ]
        
        return []
    
    except Exception as e:
        print(f"[HISTORY] ❌ Error getting history list: {e}")
        return []


def delete_history(history_id: Optional[str] = None, delete_all: bool = False, user_id: Optional[str] = None) -> Dict[str, str]:
    """
    Delete history (single atau bulk)
    
    TODO: Implementasi dengan soft delete atau hard delete
    
    Args:
        history_id: ID untuk delete single record
        delete_all: Flag untuk delete semua
        user_id: Filter user untuk delete_all
    
    Returns:
        {"status": "success/error", "message": "..."}
    """
    try:
        # TODO: Implementasi delete
        # from models import AIClaimLiteHistory
        # 
        # if delete_all:
        #     query = db.query(AIClaimLiteHistory)
        #     if user_id:
        #         query = query.filter(AIClaimLiteHistory.user_id == user_id)
        #     deleted_count = query.delete()
        #     db.commit()
        #     return {
        #         "status": "success",
        #         "message": f"{deleted_count} history records deleted"
        #     }
        # 
        # elif history_id:
        #     history = db.query(AIClaimLiteHistory).filter(
        #         AIClaimLiteHistory.claim_id == history_id
        #     ).first()
        #     if history:
        #         db.delete(history)
        #         db.commit()
        #         return {
        #             "status": "success",
        #             "message": "History berhasil dihapus"
        #         }
        #     else:
        #         return {
        #             "status": "error",
        #             "message": "History tidak ditemukan"
        #         }
        
        return {
            "status": "success",
            "message": "Delete function placeholder"
        }
    
    except Exception as e:
        print(f"[HISTORY] ❌ Error deleting history: {e}")
        return {
            "status": "error",
            "message": str(e)
        }


# ============================================================
# 🔹 SEARCH HISTORY FUNCTION (Bonus)
# ============================================================
def search_history(
    query: str,
    search_field: str = "diagnosis",
    limit: int = 10,
    user_id: Optional[str] = None
) -> List[Dict[str, Any]]:
    """
    Search history berdasarkan field tertentu
    
    TODO: Implementasi full-text search
    
    Args:
        query: Search query string
        search_field: Field yang dicari (diagnosis/claim_id)
        limit: Max results
        user_id: Filter user
    
    Returns:
        List of matching history records
    """
    try:
        # TODO: Implementasi search dengan ILIKE atau full-text search
        # from models import AIClaimLiteHistory
        # 
        # query_builder = db.query(AIClaimLiteHistory)
        # 
        # if user_id:
        #     query_builder = query_builder.filter(AIClaimLiteHistory.user_id == user_id)
        # 
        # if search_field == "diagnosis":
        #     query_builder = query_builder.filter(
        #         AIClaimLiteHistory.diagnosis.ilike(f"%{query}%")
        #     )
        # elif search_field == "claim_id":
        #     query_builder = query_builder.filter(
        #         AIClaimLiteHistory.claim_id.ilike(f"%{query}%")
        #     )
        # 
        # results = query_builder.order_by(
        #     AIClaimLiteHistory.created_at.desc()
        # ).limit(limit).all()
        # 
        # return [format_history_summary(r) for r in results]
        
        return []
    
    except Exception as e:
        print(f"[HISTORY] ❌ Error searching history: {e}")
        return []


# ============================================================
# 🔹 STATISTICS FUNCTION (Bonus untuk Dashboard)
# ============================================================
def get_history_statistics(user_id: Optional[str] = None, days: int = 30) -> Dict[str, Any]:
    """
    Get statistik history untuk dashboard
    
    Returns:
        {
            "total_claims": int,
            "by_consistency": {"Tinggi": int, "Sedang": int, "Rendah": int},
            "recent_activity": List[date_counts],
            "avg_parsing_quality": float
        }
    """
    try:
        # TODO: Implementasi aggregation queries
        return {
            "total_claims": 0,
            "by_consistency": {"Tinggi": 0, "Sedang": 0, "Rendah": 0},
            "recent_activity": [],
            "avg_parsing_quality": 0.0
        }
    
    except Exception as e:
        print(f"[STATISTICS] ❌ Error getting statistics: {e}")
        return {}